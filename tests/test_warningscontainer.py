from openpyxl import load_workbook

from customwarnings import (
    DataCheckEntityType,
    DataCheckWarning,
    DataCheckWarningLevel,
)
from warningscontainer import WarningsContainer


def test_dump_warnings_xlsx_uses_expected_column_widths_and_node_tabs(tmp_path):
    output_file = tmp_path / "warnings.xlsx"
    container = WarningsContainer()
    container.newWarning(
        DataCheckWarning(
            "CC:TypeMissing",
            "",
            "EXT",
            DataCheckWarningLevel.ERROR,
            "bbmri-eric:ID:EXT_demo:collection:col5",
            DataCheckEntityType.COLLECTION,
            "False",
            "Collection type not provided",
        )
    )
    container.newWarning(
        DataCheckWarning(
            "SE:BBDescPlaceholder",
            "",
            "SK",
            DataCheckWarningLevel.WARNING,
            "bbmri-eric:ID:SK_demo",
            DataCheckEntityType.BIOBANK,
            "False",
            "Biobank description uses a placeholder value",
        )
    )

    container.dumpWarningsXLSX(
        [str(output_file)],
        {"bbmri-eric:ID:EXT_demo": "False"},
        {"bbmri-eric:ID:EXT_demo:collection:col5": "False"},
        True,
    )

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["ALL", "AllBiobanks", "AllCollections", "EXT", "SK"]

    worksheet = workbook["ALL"]
    assert worksheet["A1"].value == "Entity ID"
    assert worksheet["D1"].value == "Check"
    assert worksheet["E1"].value == "Severity"
    assert worksheet.column_dimensions["C"].width < 10
    assert worksheet.column_dimensions["D"].width >= 24
    assert worksheet.column_dimensions["E"].width <= 12
    assert worksheet.column_dimensions["F"].width >= 100


def test_dump_suppressed_warnings_debug_logs_suppressed_entries(caplog):
    container = WarningsContainer(
        {
            "FT:KAnonViolation": {
                "bbmri-eric:ID:EU_demo:collection:col1": "reviewed false positive"
            }
        }
    )
    container.newWarning(
        DataCheckWarning(
            "FT:KAnonViolation",
            "",
            "EU",
            DataCheckWarningLevel.WARNING,
            "bbmri-eric:ID:EU_demo:collection:col1",
            DataCheckEntityType.COLLECTION,
            "False",
            "violates k-anonymity",
        )
    )

    with caplog.at_level("DEBUG"):
        container.dumpSuppressedWarningsDebug()

    assert "Suppressed warnings in this run: 1" in caplog.text
    assert "FT:KAnonViolation" in caplog.text
