import logging

from openpyxl import load_workbook
import pandas as pd
import pytest

from xlsxutils import (
    EXCEL_MAX_CELL_CHARS,
    EXCEL_TRUNCATION_MARKER,
    write_xlsx_tables,
)


def test_write_xlsx_tables_supports_multiple_sheets(tmp_path):
    output_file = tmp_path / "export.xlsx"

    write_xlsx_tables(
        str(output_file),
        [
            (pd.DataFrame([{"id": "bb1"}]), "Biobanks", False),
            (pd.DataFrame([{"id": "col1"}]), "Collections", False),
        ],
    )

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["Biobanks", "Collections"]
    assert workbook["Biobanks"]["A2"].value == "bb1"
    assert workbook["Collections"]["A2"].value == "col1"


def test_write_xlsx_tables_rejects_invalid_sheet_specs(tmp_path):
    output_file = tmp_path / "broken.xlsx"

    with pytest.raises(ValueError):
        write_xlsx_tables(str(output_file), [(pd.DataFrame(),)])


def test_write_xlsx_tables_truncates_long_text_without_mutating_input(
    tmp_path, caplog
):
    output_file = tmp_path / "long.xlsx"
    long_text = "x" * (EXCEL_MAX_CELL_CHARS + 100)
    dataframe = pd.DataFrame([{"id": "row1", "description": long_text}])
    expected_truncated_text = (
        long_text[: EXCEL_MAX_CELL_CHARS - len(EXCEL_TRUNCATION_MARKER)]
        + EXCEL_TRUNCATION_MARKER
    )

    with caplog.at_level(logging.DEBUG):
        write_xlsx_tables(str(output_file), [(dataframe, "LongText", False)])

    workbook = load_workbook(output_file, read_only=True)
    written_value = workbook["LongText"]["B2"].value
    assert len(written_value) == EXCEL_MAX_CELL_CHARS
    assert written_value.endswith(EXCEL_TRUNCATION_MARKER)
    assert dataframe.loc[0, "description"] == long_text
    assert "Truncated XLSX cells exceeding Excel's" in caplog.text
    assert "description=1" in caplog.text
    assert "Truncated XLSX cell detail: sheet=LongText" in caplog.text
    assert "entity_id=row1" in caplog.text
    assert f"original_length={len(long_text)}" in caplog.text
    assert f"original_value='{long_text}'" in caplog.text
    assert f"truncated_value='{expected_truncated_text}'" in caplog.text


def test_write_xlsx_tables_truncates_long_object_values(tmp_path, caplog):
    output_file = tmp_path / "long-object.xlsx"
    long_values = ["x" * 1000] * 40
    dataframe = pd.DataFrame([{"id": "row1", "collections": long_values}])

    with caplog.at_level(logging.INFO):
        write_xlsx_tables(str(output_file), [(dataframe, "LongObject", False)])

    workbook = load_workbook(output_file, read_only=True)
    written_value = workbook["LongObject"]["B2"].value
    assert len(written_value) == EXCEL_MAX_CELL_CHARS
    assert written_value.endswith(EXCEL_TRUNCATION_MARKER)
    assert dataframe.loc[0, "collections"] == long_values
    assert "collections=1" in caplog.text
    assert "Truncated XLSX cell detail: sheet=LongObject" in caplog.text
