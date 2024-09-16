# vim:ts=4:sw=4:tw=0:sts=4:et
"""
Script to add Orphacodes to collections in the BBMRI directory EMX file.
The script gets in input the directory EMX file, the orphanet file with mappings used by OrphaCodes class and the output
file.

For every ICD-10 diagnosis in the input file, it gets the mappings and if the orphanet mapping type is BTNT or EXACT it
adds the orphacode to the collection.

NB: The script gets the mapping Orphacodes from OrphaCodes class in orphacodes.py. The class marks the mapping type
from icd to orpha as the inverse of orpha to icd, so BTNT become NTBT. That means, that we need to check for the NTBT
mapping level in the script.
"""

import argparse
import os.path
import sys

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from orphacodes import OrphaCodes

DIAGNOSIS_AVAILABLE_COLUMN = 'diagnosis_available'
COLLECTIONS_SHEET = 'eu_bbmri_eric_collections'


def _get_diagnosis_available_column(sheet):
    for col in sheet.iter_cols(max_row=1):
        if col[0].value == DIAGNOSIS_AVAILABLE_COLUMN:
            return col[0].col_idx - 1


def add_orphacodes(directory_file, orphanet_file, output_file):
    orphacodes = OrphaCodes(orphanet_file)
    try:
        wb = openpyxl.load_workbook(directory_file)
    except InvalidFileException:
        print("{} is not a valid Directory file".format(directory_file))
        sys.exit(-1)

    try:
        sheet = wb.get_sheet_by_name(COLLECTIONS_SHEET)
    except KeyError:
        print("Missing {} sheet in file {}".format(COLLECTIONS_SHEET, directory_file))
        sys.exit(-1)

    diag_col = _get_diagnosis_available_column(sheet)
    if diag_col is None:
        print("Missing {} in sheet {}".format(DIAGNOSIS_AVAILABLE_COLUMN, directory_file))
        sys.exit(-1)

    for row in sheet.iter_rows(min_row=2):
        cell_value = row[diag_col].value
        if cell_value is not None:
            diagnosis = cell_value.split(',')
            for d in diagnosis[:]:
                orpha_codes = orphacodes.icd10ToOrpha(d.split(':')[-1])
                if orpha_codes is not None:
                    for code in orpha_codes:
                        # considering NTBT because OrphaCodes register the mapping as inverse from the Orphacode
                        if code['mapping_type'] in ('NTBT', 'E'):
                            diagnosis.append("ORPHA:{}".format(code['code']))

                    row[diag_col].value = ','.join(diagnosis)
    wb.save(output_file)


if __name__ == '__main__':
    def file_exist(file_argument):
        if os.path.exists(file_argument):
            return file_argument
        raise argparse.ArgumentTypeError("File {} does not exist".format(file_argument))

    parser = argparse.ArgumentParser()
    parser.add_argument('--directory-emx-file', '-d', dest='directory_file', type=file_exist, required=True,
                        help='Input file with list of collections and associated icd 10 codes to map')
    parser.add_argument('--orpha-mappings-file', '-O', dest='orphanet_file', type=file_exist, required=True,
                        help='XML file with orphanet icd 10/orphanet mappings')
    parser.add_argument('--output', '-o', dest='output_file', type=str, required=False,
                        help='Output file with orpacodes', default='bbmri-directory-with-orphacodes.xlsx')

    args = parser.parse_args()
    add_orphacodes(args.directory_file, args.orphanet_file, args.output_file)
